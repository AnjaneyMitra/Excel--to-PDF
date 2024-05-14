import fpdf
import pandas as pd
from fpdf import FPDF
import glob
import openpyxl
from pathlib import Path

filepaths = glob.glob(("excel data/*.xlsx"))

for filepath in filepaths:

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()
    filename = Path(filepath).stem
    invoice_no = filename.split("-")[0]
    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=12, txt=f"Invoice.nr {invoice_no}", ln=1)

    invoice_date = filename.split("-")[1]
    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=12, txt=f"Date: {invoice_date}", ln=1)

    df = pd.read_excel(filepath, sheet_name="Sheet 1")
    columns = df.columns
    columns = [item.replace("_", " ").title() for item in columns]

    pdf.set_font(family="Times", size=12, style="B")

    pdf.cell(w=30, h=12, txt=columns[0], border=1)
    pdf.cell(w=50, h=12, txt=columns[1], border=1)
    pdf.cell(w=50, h=12, txt=columns[2], border=1)
    pdf.cell(w=30, h=12, txt=columns[3], border=1)
    pdf.cell(w=30, h=12, txt=columns[4], border=1, ln=1)

    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=12, style="B")

        pdf.cell(w=30, h=12, txt=str(row["product_id"]), border=1)
        pdf.cell(w=50, h=12, txt=str(row["product_name"]), border=1)
        pdf.cell(w=50, h=12, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=12, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=12, txt=str(row["total_price"]), border=1, ln=1)

    total = df["total_price"].sum()
    pdf.cell(w=30, h=12, txt="", border=1)
    pdf.cell(w=50, h=12, txt="", border=1)
    pdf.cell(w=50, h=12, txt="", border=1)
    pdf.cell(w=30, h=12, txt="", border=1)

    pdf.cell(w=30, h=12, txt=str(total), border=1, ln=1)

    pdf.cell(w=30, h=12, txt=f"The Total sum is {total}" , ln=1)

    pdf.cell(w=30 , h=12 , txt="Anjaney Mitra")
    pdf.image("pdf.png" , w=10, h=8)
    pdf.output(f"PDFs/{filename}.pdf")
